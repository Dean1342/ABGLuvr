import httpx
import pdfplumber
import docx
import openpyxl
import csv
from io import BytesIO, StringIO


async def is_expired_discord_cdn_url(url):
    # Check if a Discord CDN image URL is expired
    if not (isinstance(url, str) and url.startswith("https://cdn.discordapp.com/")):
        return False
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.head(url, timeout=5)
            return resp.status_code >= 400
    except Exception:
        return True


async def filter_expired_images_from_content(content):
    # Remove expired Discord CDN image URLs from multimodal content
    if not isinstance(content, list):
        return content

    filtered_content = []
    for part in content:
        if isinstance(part, dict):
            if part.get("type") == "image_url":
                image_url = part.get("image_url", {}).get("url", "")
                # Only keep image if it's not expired
                if not await is_expired_discord_cdn_url(image_url):
                    filtered_content.append(part)
                # If expired, we simply skip it (don't add to filtered_content)
            else:
                # Keep non-image parts as-is
                filtered_content.append(part)
        else:
            # Keep non-dict parts as-is
            filtered_content.append(part)

    return filtered_content


async def clean_conversation_history(conversation):
    # Remove expired image URLs from conversation history
    cleaned_conversation = []

    for message in conversation:
        if isinstance(message, dict):
            cleaned_message = message.copy()

            # Check if message content contains multimodal data with images
            content = cleaned_message.get("content")
            if isinstance(content, list):
                cleaned_content = await filter_expired_images_from_content(content)
                cleaned_message["content"] = cleaned_content

            cleaned_conversation.append(cleaned_message)
        else:
            cleaned_conversation.append(message)

    return cleaned_conversation


# Multimodal content helpers
async def build_multimodal_content(message):
    # Build multimodal content from a Discord message
    content = []
    if message.reference and message.reference.resolved:
        replied = message.reference.resolved
        if replied.content:
            # Make it clearer that this is a separate entity being replied to
            user_info = (
                f"User: {replied.author.display_name} "
                f"(username: {replied.author.name}#{replied.author.discriminator}, "
                f"id: {replied.author.id})"
            )
            content.append({"type": "text", "text": f"(This message is replying to a message from {user_info}. The original message was: {replied.content})"})
        
        # Process replied message attachments
        for attachment in replied.attachments:
            if attachment.content_type and attachment.content_type.startswith("image"):
                url = attachment.url
                if not await is_expired_discord_cdn_url(url):
                    content.append({"type": "image_url", "image_url": {"url": url}})
            else:
                # Try to process as file
                file_text = await process_file_attachment(attachment)
                if file_text:
                    truncated_text = truncate_text(file_text)
                    content.append({
                        "type": "text", 
                        "text": f"Content from file '{attachment.filename}' (from replied message):\n\n{truncated_text}"
                    })
    
    if message.content:
        content.append({"type": "text", "text": message.content})
    
    # Process main message attachments
    for attachment in message.attachments:
        if attachment.content_type and attachment.content_type.startswith("image"):
            url = attachment.url
            if not await is_expired_discord_cdn_url(url):
                content.append({"type": "image_url", "image_url": {"url": url}})
        else:
            # Try to process as file
            file_text = await process_file_attachment(attachment)
            if file_text:
                truncated_text = truncate_text(file_text)
                content.append({
                    "type": "text", 
                    "text": f"Content from file '{attachment.filename}':\n\n{truncated_text}"
                })
    
    return content


async def download_file(url):
    """Download file from URL and return bytes"""
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(url, timeout=30)
            response.raise_for_status()
            return response.content
    except Exception as e:
        print(f"Error downloading file: {e}")
        return None


def extract_text_from_pdf(file_bytes):
    """Extract text from PDF bytes with better layout preservation"""
    try:
        with pdfplumber.open(BytesIO(file_bytes)) as pdf:
            text = ""
            
            for page_num, page in enumerate(pdf.pages):
                page_text = f"--- Page {page_num + 1} ---\n"
                
                # Extract tables first
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        page_text += "TABLE:\n"
                        for row in table:
                            if row:
                                page_text += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
                        page_text += "\n"
                
                # Extract regular text with better spacing
                words = page.extract_words()
                if words:
                    # Group words by lines based on y-coordinates
                    lines = {}
                    for word in words:
                        y = round(word['top'], 1)
                        if y not in lines:
                            lines[y] = []
                        lines[y].append(word)
                    
                    # Sort lines by y-coordinate and reconstruct text
                    for y in sorted(lines.keys()):
                        line_words = sorted(lines[y], key=lambda w: w['x0'])
                        line_text = ""
                        prev_x = 0
                        
                        for word in line_words:
                            # Add spacing based on x-coordinate gaps
                            gap = word['x0'] - prev_x
                            if gap > 20:  # Significant gap
                                line_text += "    "  # Add indentation
                            elif gap > 10:
                                line_text += "  "
                            line_text += word['text'] + " "
                            prev_x = word['x1']
                        
                        page_text += line_text.strip() + "\n"
                
                text += page_text + "\n"
            
            return text.strip()
    except Exception as e:
        print(f"Error extracting PDF text with pdfplumber: {e}")
        # Fallback to simple text extraction
        try:
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                text = ""
                for page_num, page in enumerate(pdf.pages):
                    page_text = f"--- Page {page_num + 1} ---\n"
                    page_text += page.extract_text() or ""
                    text += page_text + "\n"
                return text.strip()
        except Exception as e2:
            print(f"Error with fallback PDF extraction: {e2}")
            return None


def extract_text_from_docx(file_bytes):
    """Extract text from DOCX bytes"""
    try:
        doc = docx.Document(BytesIO(file_bytes))
        text = ""
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
        return text.strip()
    except Exception as e:
        print(f"Error extracting DOCX text: {e}")
        return None


def extract_text_from_xlsx(file_bytes):
    """Extract text from XLSX bytes"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes))
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
            text += "\n"
        return text.strip()
    except Exception as e:
        print(f"Error extracting XLSX text: {e}")
        return None


def extract_text_from_csv(file_bytes):
    """Extract text from CSV bytes"""
    try:
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                text_content = file_bytes.decode(encoding)
                csv_reader = csv.reader(StringIO(text_content))
                text = ""
                row_count = 0
                max_rows = 1000  # Limit to prevent huge outputs
                
                for row in csv_reader:
                    if row_count >= max_rows:
                        text += f"\n[CSV truncated after {max_rows} rows due to size limit...]"
                        break
                    
                    # Join columns with tabs for better formatting
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
                    row_count += 1
                
                return text.strip()
            except (UnicodeDecodeError, csv.Error):
                continue
        return None
    except Exception as e:
        print(f"Error extracting CSV text: {e}")
        return None


def extract_text_from_txt(file_bytes):
    """Extract text from TXT bytes"""
    try:
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                return file_bytes.decode(encoding)
            except UnicodeDecodeError:
                continue
        return None
    except Exception as e:
        print(f"Error extracting TXT text: {e}")
        return None


async def process_file_attachment(attachment):
    """Process a file attachment and extract text content"""
    if not attachment.filename:
        return None
    
    # Download the file
    file_bytes = await download_file(attachment.url)
    if not file_bytes:
        return None
    
    filename = attachment.filename.lower()
    
    # Determine file type and extract text
    if filename.endswith('.pdf'):
        return extract_text_from_pdf(file_bytes)
    elif filename.endswith('.docx'):
        return extract_text_from_docx(file_bytes)
    elif filename.endswith('.xlsx'):
        return extract_text_from_xlsx(file_bytes)
    elif filename.endswith('.csv'):
        return extract_text_from_csv(file_bytes)
    elif filename.endswith(('.txt', '.md', '.py', '.js', '.html', '.css', '.json', '.xml')):
        return extract_text_from_txt(file_bytes)
    else:
        return None


def truncate_text(text, max_chars=10000):
    """Truncate text to avoid token limits"""
    if len(text) <= max_chars:
        return text
    
    truncated = text[:max_chars]
    return truncated + "\n\n[Content truncated due to length...]"


def has_non_image_attachments(message):
    """Check if message has non-image file attachments"""
    for attachment in message.attachments:
        if not (attachment.content_type and attachment.content_type.startswith("image")):
            return True
    return False
