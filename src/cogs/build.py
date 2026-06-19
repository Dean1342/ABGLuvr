import io
import re
import json
import os
from typing import Optional

import discord
from discord.ext import commands
from discord import app_commands
import aiohttp
import openpyxl
from openai import AsyncOpenAI

from utils.integrations import supabase_client as db
from utils.ui.build_pagination import BuildPaginationView
from utils.car_charts import charts


# ── Module-level helpers ────────────────────────────────────────────────────

def _normalize_status(raw: str) -> str:
    r = str(raw).lower().strip()
    if r in ('installed', 'done', 'complete', 'completed', 'install', 'yes', '✓', '✅'):
        return 'installed'
    if r in ('ordered', 'order', 'shipping', 'en route', 'in transit', 'shipped'):
        return 'ordered'
    if r in ('in_progress', 'in progress', 'wip', 'working', 'installing'):
        return 'in_progress'
    return 'planned'


def _strip_currency(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    cleaned = re.sub(r'[^\d.]', '', str(value).replace(',', ''))
    try:
        return round(float(cleaned), 2)
    except ValueError:
        return 0.0


def _effective_cost(m: dict) -> float:
    paid = m.get('paid') or 0
    cost = m.get('cost') or 0
    return float(paid) if paid > 0 else float(cost)


def _fuzzy_match(name: str, candidates: list[str]) -> str | None:
    nl = name.lower()
    for c in candidates:
        if c.lower() == nl:
            return c
    for c in candidates:
        if nl in c.lower() or c.lower() in nl:
            return c
    return None


async def _require_profile(interaction: discord.Interaction, user_id: int) -> dict | None:
    profile = await db.get_profile(user_id)
    if not profile:
        await interaction.followup.send(
            "You don't have a car profile yet. Use `/build setcar` to set one up first.",
            ephemeral=True,
        )
    return profile


async def _download_bytes(url: str) -> bytes:
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            resp.raise_for_status()
            return await resp.read()


def _row_color_hint(row) -> str:
    """Return a text annotation if any cell in the row has a notable fill color."""
    for cell in row:
        try:
            fill = cell.fill
            if fill and fill.fill_type == 'solid':
                # Check both fgColor and bgColor — Google Sheets exports vary
                for color_obj in (fill.fgColor, fill.bgColor):
                    if color_obj.type != 'rgb':
                        continue
                    rgb = color_obj.rgb or ''
                    if len(rgb) != 8:
                        continue
                    r = int(rgb[2:4], 16)
                    g = int(rgb[4:6], 16)
                    b = int(rgb[6:8], 16)
                    # Skip transparent / black / white defaults
                    if rgb in ('00000000', 'FFFFFFFF', 'FF000000'):
                        continue
                    # Red-dominant: includes deep red AND light pink/salmon
                    if r > 150 and r > g + 30 and r > b + 30:
                        return '  [ROW HIGHLIGHT: red]'
                    # Yellow
                    if r > 200 and g > 180 and b < 100:
                        return '  [ROW HIGHLIGHT: yellow]'
                    # Green
                    if g > 160 and r < 120 and b < 120:
                        return '  [ROW HIGHLIGHT: green]'
        except Exception:
            pass
    return ''


def _xlsx_to_text(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows():
            hint = _row_color_hint(row)
            row_text = '\t'.join(str(c.value) if c.value is not None else '' for c in row)
            if row_text.strip():
                lines.append(row_text + hint)
    return '\n'.join(lines)


async def _detect_car_color(image_url: str) -> int | None:
    """Call GPT vision to identify the car's body color and return a Discord color integer."""
    try:
        client = AsyncOpenAI(api_key=os.getenv('OPENAI_API_KEY'))
        resp = await client.chat.completions.create(
            model=os.getenv('OPENAI_FINAL_MODEL', 'gpt-4.1-mini-2025-04-14'),
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": image_url}},
                    {
                        "type": "text",
                        "text": (
                            "Identify the primary body color of the car in this image. "
                            "Reply with ONLY a hex color code (#RRGGBB) representing the car's paint. "
                            "For near-black cars use #1a1a2e. For white or silver use #c0c0c0. "
                            "If no car is clearly visible, reply with #5865f2."
                        ),
                    },
                ],
            }],
            max_tokens=12,
            temperature=0,
        )
        hex_str = resp.choices[0].message.content.strip().lower()
        if not hex_str.startswith('#'):
            hex_str = '#' + hex_str
        if len(hex_str) == 7:
            r = int(hex_str[1:3], 16)
            g = int(hex_str[3:5], 16)
            b = int(hex_str[5:7], 16)
            return (r << 16) | (g << 8) | b
    except Exception:
        pass
    return None


async def _gpt_normalize_xlsx(raw_text: str) -> list[dict]:
    client = AsyncOpenAI(api_key=os.getenv('OPENAI_API_KEY'))
    system = (
        "You are a data extraction assistant. This spreadsheet is a car modification tracker. "
        "Extract every car mod — both purchased ones and planned/wishlist ones — as JSON: {\"mods\": [...]}.\n\n"
        "Each mod object:\n"
        "  name (string), "
        "category (Exterior/Interior/Performance/Suspension/Wheels & Tires/Audio/Lighting/Engine/Exhaust/Brakes/Safety/Misc), "
        "cost (number — listed or projected price), "
        "paid (number — amount actually paid; 0 if not yet purchased), "
        "status ('installed', 'planned', or 'ordered'), "
        "link (URL or null), install_date (YYYY-MM-DD or null).\n\n"
        "Determining status:\n"
        "  Purchased rows have a paid amount > 0 → 'installed' (or 'ordered' if no install date yet).\n"
        "  Planned/wishlist rows have paid = 0 and no purchase signals → 'planned'.\n"
        "  Text labels: done/complete → installed; ordered/shipping → ordered.\n"
        "  Color hints: [ROW HIGHLIGHT: red] → planned; [ROW HIGHLIGHT: green] → installed.\n\n"
        "Name markers: if a mod name contains '(OMIT)', '(SKIP)', '(WANT)', '(LATER)' or similar, "
        "this means the owner hasn't bought it yet and considers it a future mod. "
        "Include it with status='planned', paid=0, and remove the marker from the name.\n\n"
        "If separate Cost and Paid Price columns exist, map paid price → paid, listed cost → cost. "
        "Dash cams, radar detectors, GPS → Misc. Audio = speakers/amps/head units/subwoofers only. "
        "Return empty array only if no mods exist."
    )
    resp = await client.chat.completions.create(
        model=os.getenv('OPENAI_FINAL_MODEL', 'gpt-4.1-mini-2025-04-14'),
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": raw_text[:14000]},
        ],
        temperature=0,
        response_format={"type": "json_object"},
    )
    data = json.loads(resp.choices[0].message.content)
    if isinstance(data, list):
        return data
    for v in data.values():
        if isinstance(v, list):
            return v
    return []


# ── Modals ──────────────────────────────────────────────────────────────────

class EditModModal(discord.ui.Modal):
    name_field = discord.ui.TextInput(label='Name', max_length=100)
    category_field = discord.ui.TextInput(label='Category', max_length=50, required=False)
    cost_field = discord.ui.TextInput(label='Listed Cost ($)', max_length=20, required=False,
                                      placeholder='e.g. 1200 or $1,200')
    paid_field = discord.ui.TextInput(label='Amount Paid ($)', max_length=20, required=False,
                                      placeholder='e.g. 950')
    status_field = discord.ui.TextInput(
        label='Status', max_length=20, required=False,
        placeholder='installed / planned / ordered / in_progress',
    )

    def __init__(self, mod: dict):
        super().__init__(title=f"Edit: {mod['name'][:40]}")
        self.mod = mod
        self.name_field.default = mod.get('name', '')
        self.category_field.default = mod.get('category', '') or ''
        self.cost_field.default = str(mod.get('cost') or '')
        self.paid_field.default = str(mod.get('paid') or '')
        self.status_field.default = mod.get('status', 'planned')

    async def on_submit(self, interaction: discord.Interaction):
        await interaction.response.defer()
        new_name = self.name_field.value.strip() or self.mod['name']
        updated = {
            'name': new_name,
            'category': self.category_field.value.strip() or self.mod.get('category'),
            'cost': _strip_currency(self.cost_field.value) if self.cost_field.value.strip() else self.mod.get('cost'),
            'paid': _strip_currency(self.paid_field.value) if self.paid_field.value.strip() else self.mod.get('paid'),
            'status': _normalize_status(self.status_field.value) if self.status_field.value.strip() else self.mod.get('status', 'planned'),
        }
        if new_name.lower() != self.mod['name'].lower():
            await db.delete_mod(interaction.user.id, self.mod['name'])
        await db.upsert_mod(interaction.user.id, updated)
        await interaction.followup.send(f"**{new_name}** updated.")


# ── Confirmation views ──────────────────────────────────────────────────────

class ConfirmDeleteView(discord.ui.View):
    def __init__(self, mod_name: str):
        super().__init__(timeout=60)
        self.mod_name = mod_name

    @discord.ui.button(label='Confirm Delete', style=discord.ButtonStyle.danger)
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.defer()
        deleted = await db.delete_mod(interaction.user.id, self.mod_name)
        msg = f"**{self.mod_name}** removed." if deleted else "Mod not found — already removed?"
        await interaction.followup.send(msg)
        self.stop()

    @discord.ui.button(label='Cancel', style=discord.ButtonStyle.secondary)
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="Deletion cancelled.", view=None, embed=None)
        self.stop()


class ConfirmDeleteAllView(discord.ui.View):
    def __init__(self, user_id: int, count: int):
        super().__init__(timeout=60)
        self.user_id = user_id
        self.count = count

    @discord.ui.button(label='Delete All', style=discord.ButtonStyle.danger)
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.defer()
        try:
            await db.delete_all_mods(self.user_id)
            await interaction.followup.send(f"All **{self.count}** mods removed from your build.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            await interaction.followup.send(f"Failed: {e}", ephemeral=True)
        self.stop()

    @discord.ui.button(label='Cancel', style=discord.ButtonStyle.secondary)
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.edit_message(content="Cancelled.", view=None, embed=None)
        self.stop()


_IMPORT_PER_PAGE = 10


class ImportEditView(discord.ui.View):
    """Paginated import preview with inline per-mod editing."""

    def __init__(self, user_id: int, guild_id: int, mods: list[dict]):
        super().__init__(timeout=300)
        self.user_id = user_id
        self.guild_id = guild_id
        self.mods = [dict(m) for m in mods]
        self.page = 0
        self._rebuild()

    def _total_pages(self) -> int:
        return max(1, -(-len(self.mods) // _IMPORT_PER_PAGE))

    def _page_slice(self) -> list[tuple[int, dict]]:
        start = self.page * _IMPORT_PER_PAGE
        return [(start + i, m) for i, m in enumerate(self.mods[start:start + _IMPORT_PER_PAGE])]

    def _rebuild(self):
        self.clear_items()
        total = self._total_pages()
        if total > 1:
            self.add_item(_ImportNavBtn(self, 'prev', disabled=self.page <= 0, row=0))
            self.add_item(_ImportNavBtn(self, 'next', disabled=self.page >= total - 1, row=0))
        page_mods = self._page_slice()
        if page_mods:
            self.add_item(_ImportModSelect(self, page_mods))
        self.add_item(_ImportConfirmBtn(self))
        self.add_item(_ImportCancelBtn(self))

    def _make_embed(self) -> discord.Embed:
        purchased = sum(
            1 for m in self.mods
            if _normalize_status(str(m.get('status', ''))) in ('installed', 'ordered')
        )
        planned = len(self.mods) - purchased
        total_cost = sum(_effective_cost(m) for m in self.mods)
        total_pages = self._total_pages()

        lines = []
        for idx, m in self._page_slice():
            status = _normalize_status(str(m.get('status', 'planned')))
            emoji = '✅' if status == 'installed' else ('📦' if status == 'ordered' else '📋')
            cost = _effective_cost(m)
            lines.append(
                f"{emoji} **{str(m.get('name', '?'))}** "
                f"`[{str(m.get('category', 'Misc'))}]` — **${cost:,.2f}**"
            )

        emb = discord.Embed(
            title=f"Import Preview — {len(self.mods)} mods found",
            description='\n'.join(lines) or 'No mods.',
            color=discord.Color.yellow(),
        )
        emb.add_field(name='✅ Purchased', value=str(purchased), inline=True)
        emb.add_field(name='📋 Planned', value=str(planned), inline=True)
        emb.add_field(name='Total Cost', value=f"${total_cost:,.2f}", inline=True)
        emb.set_footer(
            text=f"Page {self.page + 1}/{total_pages}  •  Select a mod to edit its name, category, or status  •  Existing mods with the same name will be updated"
        )
        return emb

    async def refresh(self, interaction: discord.Interaction):
        self._rebuild()
        await interaction.response.edit_message(embed=self._make_embed(), view=self)


class _ImportNavBtn(discord.ui.Button):
    def __init__(self, view: ImportEditView, direction: str, disabled: bool, row: int):
        super().__init__(
            label='◀ Prev' if direction == 'prev' else 'Next ▶',
            style=discord.ButtonStyle.secondary,
            disabled=disabled,
            row=row,
        )
        self._v = view
        self.direction = direction

    async def callback(self, interaction: discord.Interaction):
        self._v.page += -1 if self.direction == 'prev' else 1
        await self._v.refresh(interaction)


class _ImportModSelect(discord.ui.Select):
    def __init__(self, view: ImportEditView, page_mods: list[tuple[int, dict]]):
        self._v = view
        options = []
        for idx, m in page_mods:
            status = _normalize_status(str(m.get('status', 'planned')))
            emoji = '✅' if status == 'installed' else ('📦' if status == 'ordered' else '📋')
            options.append(discord.SelectOption(
                label=str(m.get('name', '?'))[:50],
                value=str(idx),
                description=f"[{str(m.get('category', 'Misc'))[:30]}] — {status}",
                emoji=emoji,
            ))
        super().__init__(placeholder='Select a mod to edit…', options=options, row=1)

    async def callback(self, interaction: discord.Interaction):
        idx = int(self.values[0])
        await interaction.response.send_modal(_EditImportModModal(self._v, idx, self._v.mods[idx]))


class _EditImportModModal(discord.ui.Modal):
    def __init__(self, view: ImportEditView, idx: int, mod: dict):
        super().__init__(title=f"Edit: {str(mod.get('name', ''))[:40]}")
        self._v = view
        self.idx = idx
        cost = _effective_cost(mod)
        self.name_field = discord.ui.TextInput(
            label='Name', max_length=100, default=str(mod.get('name', ''))
        )
        self.category_field = discord.ui.TextInput(
            label='Category', max_length=50, required=False,
            placeholder='Exterior / Performance / Audio / Misc / etc.',
            default=str(mod.get('category', '')) or '',
        )
        self.cost_field = discord.ui.TextInput(
            label='Cost ($)', max_length=20, required=False,
            default=f'{cost:.2f}' if cost else '',
        )
        self.status_field = discord.ui.TextInput(
            label='Status', max_length=20, required=False,
            placeholder='installed / planned / ordered',
            default=str(mod.get('status', 'planned')),
        )
        for item in (self.name_field, self.category_field, self.cost_field, self.status_field):
            self.add_item(item)

    async def on_submit(self, interaction: discord.Interaction):
        mod = self._v.mods[self.idx]
        mod['name'] = self.name_field.value.strip() or mod.get('name', 'Unknown')
        mod['category'] = self.category_field.value.strip() or mod.get('category', 'Misc')
        mod['status'] = (
            _normalize_status(self.status_field.value)
            if self.status_field.value.strip()
            else mod.get('status', 'planned')
        )
        if self.cost_field.value.strip():
            v = _strip_currency(self.cost_field.value)
            mod['paid'] = v
            mod['cost'] = v
        self._v._rebuild()
        await interaction.response.edit_message(embed=self._v._make_embed(), view=self._v)


class _ImportConfirmBtn(discord.ui.Button):
    def __init__(self, view: ImportEditView):
        super().__init__(label='Confirm Import', style=discord.ButtonStyle.success, emoji='✅', row=2)
        self._v = view

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.defer()
        try:
            normalized = [
                {
                    'name': str(m.get('name', 'Unknown'))[:100],
                    'category': str(m.get('category', 'Misc'))[:50],
                    'cost': _strip_currency(m.get('cost')),
                    'paid': _strip_currency(m.get('paid')),
                    'status': _normalize_status(str(m.get('status', 'planned'))),
                    'link': m.get('link') or None,
                    'install_date': m.get('install_date') or None,
                }
                for m in self._v.mods
            ]
            await db.bulk_upsert_mods(self._v.user_id, normalized)
            await interaction.followup.send(f"Imported **{len(normalized)}** mods to your build.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            await interaction.followup.send(f"Import failed: {e}", ephemeral=True)
        self._v.stop()


class _ImportCancelBtn(discord.ui.Button):
    def __init__(self, view: ImportEditView):
        super().__init__(label='Cancel', style=discord.ButtonStyle.danger, emoji='❌', row=2)
        self._v = view

    async def callback(self, interaction: discord.Interaction):
        await interaction.response.edit_message(content="Import cancelled.", embed=None, view=None)
        self._v.stop()


# ── Help system ─────────────────────────────────────────────────────────────

def _make_overview_embed() -> discord.Embed:
    emb = discord.Embed(
        title='🚗  Build Tracker — Help',
        description=(
            'Track your car mods, costs, and build progress right inside Discord.\n'
            'Data is saved per-user and persists across bot restarts.\n\n'
            'Use the dropdown below to see detailed usage for any command.'
        ),
        color=discord.Color.blurple(),
    )
    emb.add_field(
        name='⚙️  Profile Setup',
        value=(
            '`/build setcar` — Set your car *(required first step)*\n'
            '`/build setbio` — Write a bio and choose your embed color\n'
            '`/build setimage` — Add or remove car photos (main + thumbnail)'
        ),
        inline=True,
    )
    emb.add_field(
        name='🔧  Mod Tracking',
        value=(
            '`/build add` — Add a part to your build\n'
            '`/build remove` — Remove a part *(with confirmation)*\n'
            '`/build edit` — Edit a part via a pre-filled form'
        ),
        inline=True,
    )
    emb.add_field(name='​', value='​', inline=True)
    emb.add_field(
        name='📊  Views & Analysis',
        value='`/build view [@user]` — Full paginated build with charts',
        inline=True,
    )
    emb.add_field(
        name='💰  Labor Costs',
        value=(
            '`/build labor add` — Log a shop or install cost\n'
            '`/build labor remove` — Remove a labor entry'
        ),
        inline=True,
    )
    emb.add_field(
        name='📥  Import',
        value='`/build import` — Import mods from an Excel spreadsheet',
        inline=True,
    )
    emb.set_footer(text='Select a command from the dropdown for detailed usage ↓')
    return emb


def _make_command_embed(key: str) -> discord.Embed:
    if key == '__overview__':
        return _make_overview_embed()

    e = discord.Embed(color=discord.Color.blurple())

    if key == 'setcar':
        e.title = '/build setcar — Set Your Car'
        e.description = (
            'Register your car\'s year, make, and model. **Required before using most other commands.** '
            'Running it again updates your info without touching your mods.'
        )
        e.add_field(
            name='Parameters',
            value='`year` — Model year *(e.g. 2022)*\n`make` — Brand *(e.g. Subaru)*\n`model` — Model name *(e.g. WRX)*',
            inline=False,
        )
        e.add_field(name='Usage', value='`/build setcar 2022 Subaru WRX`', inline=False)

    elif key == 'setbio':
        e.title = '/build setbio — Set Bio & Embed Color'
        e.description = (
            'Update your build bio and/or choose the color used for all your build embeds. '
            'All parameters are optional — you can update just the bio, just the color, or both at once.'
        )
        e.add_field(
            name='Parameters (all optional)',
            value=(
                '`text` — Bio shown on your profile *(max 500 characters)*\n'
                '`color` — Pick from a preset list of colors\n'
                '`customcolor` — Enter any hex code *(e.g. `#FF5500`)* — overrides the preset'
            ),
            inline=False,
        )
        e.add_field(
            name='Usage',
            value=(
                '`/build setbio text:Daily driven, targeting 350whp`\n'
                '`/build setbio color:Red`\n'
                '`/build setbio customcolor:#FF5500`\n'
                '`/build setbio text:Stage 2 complete. color:Green`'
            ),
            inline=False,
        )
        e.add_field(
            name='Color Presets',
            value='Default (Blurple) · Red · Orange · Yellow · Green · Teal · Blue · Purple · Hot Pink · White · Dark',
            inline=False,
        )

    elif key == 'setimage':
        e.title = '/build setimage — Set or Remove Car Photos'
        e.description = (
            'Add or remove photos on your build profile. '
            'Set via URL or file upload. The main image also auto-detects your car\'s color for the embed theme.'
        )
        e.add_field(
            name='Setting photos (all optional — provide any combination)',
            value=(
                '`main` — URL for the large embed image\n'
                '`thumb` — URL for the small top-right thumbnail\n'
                '`mainfile` — Upload the main image from your device\n'
                '`thumbfile` — Upload the thumbnail from your device'
            ),
            inline=False,
        )
        e.add_field(
            name='Removing photos',
            value=(
                '`remove:Main image` — Clears the main embed image\n'
                '`remove:Thumbnail` — Clears the thumbnail\n'
                '`remove:All photos` — Clears both'
            ),
            inline=False,
        )
        e.add_field(
            name='Usage',
            value=(
                '`/build setimage main:https://i.imgur.com/abc.jpg`\n'
                '`/build setimage mainfile:[attach photo]`  *(upload from device)*\n'
                '`/build setimage remove:All photos`  *(remove everything)*'
            ),
            inline=False,
        )
        e.add_field(
            name='Notes',
            value=(
                '• File uploads take priority over URLs if both are provided for the same slot\n'
                '• Setting a main image automatically detects your car\'s paint color and applies it as the embed theme\n'
                '• Override the auto-detected color any time with `/build setbio customcolor:#RRGGBB`'
            ),
            inline=False,
        )

    elif key == 'add':
        e.title = '/build add — Add a Mod'
        e.description = (
            'Add a part or modification to your build tracker. '
            'Running this with the same mod name updates the existing entry instead of duplicating it.'
        )
        e.add_field(
            name='Required',
            value=(
                '`name` — Part name *(e.g. "Coilover Kit")*\n'
                '`category` — e.g. Exterior / Performance / Interior / Wheels & Tires / Audio\n'
                '`cost` — Listed or total price *(e.g. 2500 or $2,500)*\n'
                '`paid` — Amount you\'ve actually paid so far'
            ),
            inline=False,
        )
        e.add_field(
            name='Optional',
            value=(
                '`link` — Product page or receipt URL\n'
                '`status` — `planned` / `ordered` / `installed` / `in_progress`  ·  default: planned\n'
                '`date` — Install date in `YYYY-MM-DD` format'
            ),
            inline=False,
        )
        e.add_field(
            name='Usage',
            value='`/build add "Whiteline Sway Bar" Suspension 280 280 installed 2024-03-10`',
            inline=False,
        )

    elif key == 'remove':
        e.title = '/build remove — Remove a Mod'
        e.description = (
            'Remove a mod from your build tracker. '
            'A confirmation embed appears first — nothing is deleted until you confirm.'
        )
        e.add_field(name='Parameters', value='`name` — Mod name *(fuzzy matched, doesn\'t need to be exact)*', inline=False)
        e.add_field(
            name='Usage',
            value='`/build remove "Coilover Kit"`\n`/build remove coilover`  *(partial match works)*',
            inline=False,
        )
        e.add_field(name='Tips', value='Use `/build view` to see your exact mod names if the fuzzy match misses.', inline=False)

    elif key == 'edit':
        e.title = '/build edit — Edit a Mod'
        e.description = 'Opens a pre-filled form to update an existing mod\'s details. Leave any field blank to keep its current value.'
        e.add_field(name='Parameters', value='`name` — Mod name to edit *(fuzzy matched)*', inline=False)
        e.add_field(
            name='Form Fields',
            value='Name · Category · Listed Cost · Amount Paid · Status',
            inline=False,
        )
        e.add_field(name='Usage', value='`/build edit "Coilover Kit"`', inline=False)
        e.add_field(
            name='Notes',
            value='Renaming a mod (changing the Name field) is supported — the old entry is replaced.',
            inline=False,
        )

    elif key == 'view':
        e.title = '/build view — View a Build'
        e.description = (
            'Full paginated embed view of a build — profile, mod list, and automatically generated charts. '
            'Navigate pages with the ◀️ ▶️ buttons.'
        )
        e.add_field(name='Parameters', value='`user` — *(optional)* @mention to view someone else\'s build. Defaults to your own.', inline=False)
        e.add_field(
            name='Pages',
            value=(
                '**1️⃣ Profile** — Car, bio, images, cost summary stats\n'
                '**2️⃣+ Mods** — Full list with status, costs, and links *(8 per page)*\n'
                '**📊 Status Donut** — Installed vs Planned vs Ordered breakdown\n'
                '**📊 Budget Bar** — How much of your projected build you\'ve spent\n'
                '**📊 Timeline** — Install dates plotted over time *(if dates are logged)*\n'
                '**📊 Categories** — Spend broken down by category'
            ),
            inline=False,
        )
        e.add_field(name='Usage', value='`/build view`\n`/build view @friend`', inline=False)

    elif key == 'import':
        e.title = '/build import — Import from Excel'
        e.description = (
            'Reads your existing Excel spreadsheet and imports all mods automatically. '
            'GPT normalizes any column layout so it works regardless of how your sheet is structured.'
        )
        e.add_field(name='Parameters', value='`attachment` — Attach your `.xlsx` file when running the command', inline=False)
        e.add_field(
            name='How It Works',
            value=(
                '1. Run `/build import` and attach your `.xlsx` file\n'
                '2. GPT scans every sheet and extracts mods regardless of column order\n'
                '3. A preview shows what was found — counts, totals, categories\n'
                '4. Hit **Confirm ✅** to save or **Cancel ❌** to discard'
            ),
            inline=False,
        )
        e.add_field(
            name='What It Handles',
            value=(
                '• Any column layout or sheet name\n'
                '• Status inferred from text labels *(done, installed, want, etc.)*\n'
                '• Prices with $, commas, or other formatting\n'
                '• Duplicate mod names are updated, not doubled'
            ),
            inline=False,
        )

    elif key == 'labor_add':
        e.title = '/build labor add — Add Labor Cost'
        e.description = (
            'Track shop visits, installation fees, or any paid work on your car. '
            'Labor costs appear separately in your summary and are included in the grand total.'
        )
        e.add_field(
            name='Parameters',
            value=(
                '`description` — What was done *(e.g. "Alignment", "Full suspension install")*\n'
                '`cost` — Cost in dollars\n'
                '`date` — *(optional)* Date the work was done `YYYY-MM-DD`'
            ),
            inline=False,
        )
        e.add_field(
            name='Usage',
            value='`/build labor add "Alignment" 150`\n`/build labor add "Suspension install" 800 2024-02-20`',
            inline=False,
        )

    elif key == 'labor_remove':
        e.title = '/build labor remove — Remove Labor Entry'
        e.description = 'Remove a labor cost entry by matching its description. Partial matches are supported.'
        e.add_field(name='Parameters', value='`description` — Partial description match *(doesn\'t need to be exact)*', inline=False)
        e.add_field(
            name='Usage',
            value='`/build labor remove "Alignment"`\n`/build labor remove install`  *(matches any entry containing "install")*',
            inline=False,
        )
        e.add_field(name='Notes', value='If multiple entries match, the bot lists them so you can be more specific.', inline=False)

    else:
        e.title = 'Unknown Command'
        e.description = 'Select a command from the dropdown below.'

    e.set_footer(text='Select another command from the dropdown ↓')
    return e


class BuildCommandSelect(discord.ui.Select):
    def __init__(self):
        options = [
            discord.SelectOption(label='← Overview', value='__overview__', description='Back to the full command list'),
            discord.SelectOption(label='/build setcar', value='setcar', emoji='🚗', description='Set your car year, make, and model'),
            discord.SelectOption(label='/build setbio', value='setbio', emoji='📝', description='Write a bio for your build profile'),
            discord.SelectOption(label='/build setimage', value='setimage', emoji='🖼️', description='Add car photos (main image + thumbnail)'),
            discord.SelectOption(label='/build add', value='add', emoji='➕', description='Add a part or mod to your build'),
            discord.SelectOption(label='/build remove', value='remove', emoji='🗑️', description='Remove a mod (shows confirmation first)'),
            discord.SelectOption(label='/build edit', value='edit', emoji='✏️', description='Edit a mod via a pre-filled form'),
            discord.SelectOption(label='/build view', value='view', emoji='👁️', description='Full paginated build view with charts'),
            discord.SelectOption(label='/build import', value='import', emoji='📥', description='Import mods from an Excel spreadsheet'),
            discord.SelectOption(label='/build labor add', value='labor_add', emoji='🔧', description='Log a shop or installation cost'),
            discord.SelectOption(label='/build labor remove', value='labor_remove', emoji='❌', description='Remove a labor cost entry'),
        ]
        super().__init__(
            placeholder='Select a command for detailed usage…',
            options=options,
            min_values=1,
            max_values=1,
        )

    async def callback(self, interaction: discord.Interaction):
        embed = _make_command_embed(self.values[0])
        await interaction.response.edit_message(embed=embed, view=self.view)


# ── Labor group ─────────────────────────────────────────────────────────────

labor_group = app_commands.Group(name="labor", description="Track labor costs for your build.")


@labor_group.command(name="add", description="Add a labor cost entry.")
@app_commands.describe(
    description="What work was done",
    cost="Labor cost in dollars",
    date="Date of work (YYYY-MM-DD, optional)",
)
async def labor_add(
    interaction: discord.Interaction,
    description: str,
    cost: str,
    date: Optional[str] = None,
):
    await interaction.response.defer()
    profile = await db.get_profile(interaction.user.id)
    if not profile:
        await interaction.followup.send(
            "You don't have a car profile yet. Use `/build setcar` first.", ephemeral=True
        )
        return
    await db.insert_labor(interaction.user.id, {
        'description': description[:200],
        'cost': _strip_currency(cost),
        'date': date or None,
    })
    await interaction.followup.send(
        f"Labor added: **{description}** — ${_strip_currency(cost):,.2f}"
    )


@labor_group.command(name="remove", description="Remove a labor cost entry.")
@app_commands.describe(description="Description of the labor entry to remove")
async def labor_remove(interaction: discord.Interaction, description: str):
    await interaction.response.defer()
    labor = await db.get_labor(interaction.user.id)
    if not labor:
        await interaction.followup.send("No labor entries found.", ephemeral=True)
        return
    matches = [l for l in labor if description.lower() in l['description'].lower()]
    if not matches:
        await interaction.followup.send(
            f"No labor entry found matching **{description}**.", ephemeral=True
        )
        return
    if len(matches) == 1:
        await db.delete_labor(interaction.user.id, matches[0]['id'])
        await interaction.followup.send(f"Removed: **{matches[0]['description']}**")
    else:
        lines = [
            f"{i+1}. {l['description']} — ${(l.get('cost') or 0):,.2f} ({l.get('date', 'no date')})"
            for i, l in enumerate(matches)
        ]
        await interaction.followup.send(
            "Multiple matches found — be more specific:\n" + '\n'.join(lines), ephemeral=True
        )


# ── Main cog ────────────────────────────────────────────────────────────────

class Build(commands.GroupCog, name="build"):
    """Per-user car mod build tracker."""

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    # Attach the labor sub-group so it becomes /build labor add / remove
    labor = labor_group

    # ── Profile setup ─────────────────────────────────────────────────────

    @app_commands.command(name="setcar", description="Set your car (year, make, model).")
    @app_commands.describe(year="Model year (e.g. 2022)", make="Brand (e.g. Subaru)", model="Model (e.g. WRX)")
    async def setcar(self, interaction: discord.Interaction, year: int, make: str, model: str):
        await interaction.response.defer()
        await db.upsert_profile(interaction.user.id, interaction.guild_id or 0, {
            'year': year, 'make': make, 'model': model,
        })
        await interaction.followup.send(f"{interaction.user.mention} set their car to **{year} {make} {model}**.")

    @app_commands.command(name="setbio", description="Set your build bio and/or embed color.")
    @app_commands.describe(
        text="Bio shown on your build profile (max 500 chars)",
        color="Preset embed color for your build pages",
        customcolor="Custom hex color code — overrides preset (e.g. FF5500 or #FF5500)",
    )
    @app_commands.choices(color=[
        app_commands.Choice(name='Default (Blurple)', value='5865f2'),
        app_commands.Choice(name='Red',               value='ed4245'),
        app_commands.Choice(name='Orange',            value='e67e22'),
        app_commands.Choice(name='Yellow',            value='fee75c'),
        app_commands.Choice(name='Green',             value='57f287'),
        app_commands.Choice(name='Teal',              value='1abc9c'),
        app_commands.Choice(name='Blue',              value='3498db'),
        app_commands.Choice(name='Purple',            value='9b59b6'),
        app_commands.Choice(name='Hot Pink',          value='eb459e'),
        app_commands.Choice(name='White',             value='f2f3f5'),
        app_commands.Choice(name='Dark',              value='1a1a2e'),
    ])
    async def setbio(
        self, interaction: discord.Interaction,
        text: Optional[str] = None,
        color: Optional[str] = None,
        customcolor: Optional[str] = None,
    ):
        await interaction.response.defer()

        if text and len(text) > 500:
            await interaction.followup.send("Bio must be 500 characters or fewer.", ephemeral=True)
            return

        embed_color: int | None = None
        if customcolor:
            hex_str = customcolor.strip().lstrip('#')
            try:
                embed_color = int(hex_str, 16) & 0xFFFFFF
            except ValueError:
                await interaction.followup.send(
                    "Invalid hex color — use a format like `#FF5500` or `FF5500`.", ephemeral=True
                )
                return
        elif color:
            embed_color = int(color, 16)

        updates: dict = {}
        if text:
            updates['bio'] = text
        if embed_color is not None:
            updates['embed_color'] = embed_color

        if not updates:
            await interaction.followup.send(
                "Provide a bio, a color, or both.", ephemeral=True
            )
            return

        await db.upsert_profile(interaction.user.id, interaction.guild_id or 0, updates)

        parts = []
        if text:
            parts.append('bio')
        if embed_color is not None:
            parts.append(f"embed color (`#{embed_color:06x}`)")
        await interaction.followup.send(
            f"{interaction.user.mention} updated their build {' and '.join(parts)}."
        )

    @app_commands.command(name="setimage", description="Set or remove your car photos.")
    @app_commands.describe(
        main="URL for the large embed image",
        thumb="URL for the small top-right thumbnail",
        mainfile="Upload the main image from your device",
        thumbfile="Upload the thumbnail from your device",
        remove="Remove photo(s) from your profile",
    )
    @app_commands.choices(remove=[
        app_commands.Choice(name='Main image',  value='main'),
        app_commands.Choice(name='Thumbnail',   value='thumb'),
        app_commands.Choice(name='All photos',  value='all'),
    ])
    async def setimage(
        self, interaction: discord.Interaction,
        main: Optional[str] = None,
        thumb: Optional[str] = None,
        mainfile: Optional[discord.Attachment] = None,
        thumbfile: Optional[discord.Attachment] = None,
        remove: Optional[str] = None,
    ):
        await interaction.response.defer()

        if remove:
            clear: dict = {}
            if remove in ('main', 'all'):
                clear['car_image_url'] = None
            if remove in ('thumb', 'all'):
                clear['thumbnail_url'] = None
            await db.upsert_profile(interaction.user.id, interaction.guild_id or 0, clear)
            label = {'main': 'main image', 'thumb': 'thumbnail', 'all': 'all photos'}[remove]
            await interaction.followup.send(
                f"{interaction.user.mention} removed their {label}."
            )
            return

        updates: dict = {}
        labels: list[str] = []

        main_final = mainfile.url if mainfile else main
        if main_final:
            updates['car_image_url'] = main_final
            labels.append('main image')

        thumb_final = thumbfile.url if thumbfile else thumb
        if thumb_final:
            updates['thumbnail_url'] = thumb_final
            labels.append('thumbnail')

        if not updates:
            await interaction.followup.send(
                "Provide at least one image (URL or file), or use `remove` to clear photos.",
                ephemeral=True,
            )
            return

        await db.upsert_profile(interaction.user.id, interaction.guild_id or 0, updates)

        if main_final:
            detected = await _detect_car_color(main_final)
            if detected is not None:
                await db.upsert_profile(
                    interaction.user.id, interaction.guild_id or 0, {'embed_color': detected}
                )

        await interaction.followup.send(
            f"{interaction.user.mention} updated their build {' and '.join(labels)}."
        )

    # ── Mod commands ──────────────────────────────────────────────────────

    @app_commands.command(name="add", description="Add a mod to your build.")
    @app_commands.describe(
        name="Part/mod name",
        category="Category (Exterior, Performance, Interior, Wheels & Tires, Audio, etc.)",
        cost="Listed price (e.g. 1200 or $1,200.00)",
        paid="Amount you've paid so far",
        link="Product page or receipt URL (optional)",
        status="Current status",
        date="Install date YYYY-MM-DD (optional)",
    )
    @app_commands.choices(status=[
        app_commands.Choice(name='Planned', value='planned'),
        app_commands.Choice(name='Ordered', value='ordered'),
        app_commands.Choice(name='Installed', value='installed'),
        app_commands.Choice(name='In Progress', value='in_progress'),
    ])
    async def add(
        self, interaction: discord.Interaction,
        name: str,
        category: str,
        cost: str,
        paid: str,
        link: Optional[str] = None,
        status: str = 'planned',
        date: Optional[str] = None,
    ):
        await interaction.response.defer()
        profile = await _require_profile(interaction, interaction.user.id)
        if not profile:
            return
        mod_data = {
            'name': name[:100],
            'category': category[:50],
            'cost': _strip_currency(cost),
            'paid': _strip_currency(paid),
            'status': _normalize_status(status),
            'link': link or None,
            'install_date': date or None,
        }
        await db.upsert_mod(interaction.user.id, mod_data)
        await interaction.followup.send(
            f"{interaction.user.mention} added **{name}** `[{category}]` — "
            f"${mod_data['cost']:,.2f} listed, ${mod_data['paid']:,.2f} paid  ({mod_data['status']})"
        )

    @app_commands.command(name="remove", description="Remove a mod from your build, or remove all mods.")
    @app_commands.describe(
        name="Mod name to remove (fuzzy matched)",
        removeall="Remove ALL mods from your build (shows confirmation first)",
    )
    async def remove(
        self, interaction: discord.Interaction,
        name: Optional[str] = None,
        removeall: bool = False,
    ):
        await interaction.response.defer()
        profile = await _require_profile(interaction, interaction.user.id)
        if not profile:
            return

        if removeall:
            mods = await db.get_mods(interaction.user.id)
            if not mods:
                await interaction.followup.send("You have no mods to remove.", ephemeral=True)
                return
            emb = discord.Embed(
                title=f"Remove ALL {len(mods)} mods?",
                description="This will permanently delete your entire mod list. This cannot be undone.",
                color=discord.Color.red(),
            )
            await interaction.followup.send(embed=emb, view=ConfirmDeleteAllView(interaction.user.id, len(mods)))
            return

        if not name:
            await interaction.followup.send(
                "Provide a mod name, or set `remove_all: True` to clear your entire list.", ephemeral=True
            )
            return

        mods = await db.get_mods(interaction.user.id)
        match = _fuzzy_match(name, [m['name'] for m in mods])
        if not match:
            await interaction.followup.send(
                f"No mod found matching **{name}**. Check `/build view` for your list.", ephemeral=True
            )
            return
        mod = next(m for m in mods if m['name'] == match)
        emb = discord.Embed(
            title=f"Remove **{match}**?",
            description=(
                f"Category: {mod.get('category', '—')}\n"
                f"Cost: ${_effective_cost(mod):,.2f}\n"
                f"Status: {mod.get('status', '—')}"
            ),
            color=discord.Color.red(),
        )
        view = ConfirmDeleteView(match)
        await interaction.followup.send(embed=emb, view=view)

    @app_commands.command(name="edit", description="Edit an existing mod's details.")
    @app_commands.describe(name="Mod name to edit (fuzzy matched)")
    async def edit(self, interaction: discord.Interaction, name: str):
        # Cannot defer — Modal must be the initial response
        mods = await db.get_mods(interaction.user.id)
        if not mods:
            await interaction.response.send_message(
                "You have no mods to edit. Use `/build add` first.", ephemeral=True
            )
            return
        match = _fuzzy_match(name, [m['name'] for m in mods])
        if not match:
            await interaction.response.send_message(
                f"No mod found matching **{name}**.", ephemeral=True
            )
            return
        mod = next(m for m in mods if m['name'] == match)
        await interaction.response.send_modal(EditModModal(mod))

    # ── View / Summary ────────────────────────────────────────────────────

    @app_commands.command(name="view", description="View a car build with charts and mod list.")
    @app_commands.describe(user="User to view (defaults to yourself)")
    async def view(self, interaction: discord.Interaction, user: Optional[discord.User] = None):
        await interaction.response.defer()
        target = user or interaction.user
        profile = await db.get_profile(target.id)
        if not profile:
            msg = (
                "You haven't set up a build yet. Use `/build setcar` to get started."
                if user is None else
                f"{target.display_name} hasn't set up their build yet."
            )
            await interaction.followup.send(msg, ephemeral=True)
            return

        mods = await db.get_mods(target.id)
        labor = await db.get_labor(target.id)

        installed = sum(1 for m in mods if m.get('status') == 'installed')
        ordered = sum(1 for m in mods if m.get('status') == 'ordered')
        planned = sum(1 for m in mods if m.get('status') == 'planned')
        total_paid = sum(m.get('paid') or 0 for m in mods)
        total_cost = sum(m.get('cost') or 0 for m in mods)
        category_totals: dict[str, float] = {}
        for m in mods:
            cat = m.get('category') or 'Misc'
            category_totals[cat] = category_totals.get(cat, 0) + (m.get('paid') or 0)

        chart_donut = charts.generate_donut_chart(installed, planned, ordered) if mods else None
        chart_budget = charts.generate_budget_chart(total_paid, total_cost) if total_cost > 0 else None
        chart_timeline = charts.generate_timeline_chart(mods) if mods else None
        chart_categories = charts.generate_category_chart(category_totals) if category_totals else None

        view = BuildPaginationView(
            profile=profile, mods=mods, labor=labor, target_user=target,
            chart_donut=chart_donut, chart_budget=chart_budget,
            chart_timeline=chart_timeline, chart_categories=chart_categories,
        )
        embed, file = view.get_current_embed_and_file()
        if file:
            await interaction.followup.send(embed=embed, view=view, file=file)
        else:
            await interaction.followup.send(embed=embed, view=view)

    # ── Help ──────────────────────────────────────────────────────────────

    @app_commands.command(name="help", description="Overview and usage guide for all /build commands.")
    async def help_cmd(self, interaction: discord.Interaction):
        view = discord.ui.View(timeout=300)
        view.add_item(BuildCommandSelect())
        await interaction.response.send_message(embed=_make_overview_embed(), view=view)

    # ── Import ────────────────────────────────────────────────────────────

    @app_commands.command(name="import", description="Import mods from an Excel (.xlsx) spreadsheet.")
    @app_commands.describe(attachment="Your .xlsx file — all sheets will be scanned")
    async def import_build(self, interaction: discord.Interaction, attachment: discord.Attachment):
        await interaction.response.defer()
        profile = await _require_profile(interaction, interaction.user.id)
        if not profile:
            return

        if not attachment.filename.lower().endswith('.xlsx'):
            await interaction.followup.send("Please upload an `.xlsx` file.", ephemeral=True)
            return

        try:
            file_bytes = await _download_bytes(attachment.url)
        except Exception as e:
            await interaction.followup.send(f"Failed to download the file: {e}", ephemeral=True)
            return

        try:
            raw_text = _xlsx_to_text(file_bytes)
        except Exception as e:
            await interaction.followup.send(f"Failed to read the spreadsheet: {e}", ephemeral=True)
            return

        await interaction.edit_original_response(content="Reading your spreadsheet with GPT…")

        try:
            mods = await _gpt_normalize_xlsx(raw_text)
        except Exception as e:
            await interaction.edit_original_response(content=f"GPT extraction failed: {e}")
            return

        if not mods:
            await interaction.edit_original_response(
                content="No mods could be extracted. Make sure your sheet has part names and prices."
            )
            return

        view = ImportEditView(interaction.user.id, interaction.guild_id or 0, mods)
        await interaction.edit_original_response(content="", embed=view._make_embed(), view=view)


async def setup(bot: commands.Bot):
    await bot.add_cog(Build(bot))
